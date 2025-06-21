import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl.styles import Alignment

class Training_Results():
    frame_idx = None
    episode_reward = None
    value_loss = None
    policy_loss = None

def generate_fake_data():
    n_points = 200
    frame_idxs = np.arange(1, n_points + 1)
    rewards = np.random.uniform(-500, 500, n_points)
    value_losses = np.random.uniform(10000, 50000, n_points)
    policy_losses = np.random.uniform(100, 500, n_points)

    results = {
        "frame_idx": frame_idxs.tolist(),
        "episode_reward": rewards.tolist(),
        "value_loss": value_losses.tolist(),
        "policy_loss": policy_losses.tolist()
    }
    return results

def dic_to_training_results(dic):
    results = Training_Results()
    for element in dic:
        setattr(results, element, dic[element].tolist())
    return results

def training_results_to_dic(data):
    return data.__dict__

def save_result_to_excel(data, file_path):
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max_length + 2

def load_result(file_path):
    results = Training_Results()
    df = pd.read_csv(file_path)

    for column in df.columns:
        setattr(results, column, df[column].tolist())
    return results

def save_result(frame_idxs, frame_rewards, value_losses, policy_losses, results_folder, frame_controller_parrams, reward_list_record):
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 1

    params_names = ["Kpp", "Kvp", "Kvi"]
    action_names = ["ΔKpp", "ΔKvp", "ΔKvi"]

    for iteration_num, iteration_data in enumerate(frame_controller_parrams, start=1):
        iteration_title = f"Iteration {iteration_num}"
        ws.cell(row=row, column=1, value=iteration_title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center') 
        row += 1

        controller_title = f"Controller"
        ws.cell(row=row, column=1, value=controller_title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        action_title = f"Action"
        ws.cell(row=row, column=5, value=action_title)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center') 

        for idx, name in enumerate(params_names, start=1):
            ws.cell(row=row+1, column=idx, value=name)
            ws.cell(row=row+1, column=idx).alignment = Alignment(horizontal='center') 
        for idx, name in enumerate(action_names, start=1):
            ws.cell(row=row+1, column=idx+4, value=name)
            ws.cell(row=row+1, column=idx+4).alignment = Alignment(horizontal='center') 
        
        row += 2
        for step_num, step_params in enumerate(iteration_data, start=1):
            for col_num, param in enumerate(step_params, start=1):
                ws.cell(row=row, column=col_num, value=param)

            if step_num > 1:
                for col_num, param in enumerate(step_params, start=1):
                    previous_step_params = iteration_data[step_num - 2]
                    difference = param - previous_step_params[col_num - 1]
                    ws.cell(row=row-1, column=col_num + 4, value=difference)

            row += 1
        row += 1

    wb.save(results_folder + "iterations_params.xlsx")

    ## Reward
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 1
    
    reward_names = [
            "overshoot", "overshoot_load", "settling_time", "settling_time_load", \
            "max_torque_flag", "GM_velocity",\
            "Emax_before", "Eavg_before", "Emax_after", "Eavg_after",\
            "settling_time_001p", "settling_time_01p", "settling_time_1p", "settling_time_3p", \
            "damping_ratio", "Ess_step", "Ess_trap"
        ]

    for iteration_num, iteration_data in enumerate(reward_list_record, start=1):
        iteration_title = f"Iteration {iteration_num}"
        ws.cell(row=row, column=1, value=iteration_title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center') 

        for idx, name in enumerate(reward_names, start=1):
            ws.cell(row=row+1, column=idx, value=name)
            ws.cell(row=row+1, column=idx).alignment = Alignment(horizontal='center') 
        
        row += 2
        for step_num, step_params in enumerate(iteration_data, start=1):
            for col_num, param in enumerate(step_params, start=1):
                ws.cell(row=row, column=col_num, value=param)
            row += 1
        row += 1
    wb.save(results_folder + "actual_reward.xlsx")

    frame_rewards = np.array(frame_rewards)
    graph_folder = results_folder + 'graph/'
    os.makedirs(graph_folder, exist_ok=True)

    # "overshoot", "settling_time", "Ess", "Emax_before", 
    # "Emax_after", "Eavg_before", "Eavg_after", "max_torque", "total_reward"
    reward_info = {
        "frame_idx"                 : frame_idxs,
        "frame_overshoot"           : frame_rewards[:, 0].tolist(),
        "frame_settling_time"       : frame_rewards[:, 1].tolist(),
        "frame_Emax_before"         : frame_rewards[:, 2].tolist(),
        "frame_Emax_after"          : frame_rewards[:, 3].tolist(),
        "frame_Eavg_before"         : frame_rewards[:, 4].tolist(),
        "frame_Eavg_after"          : frame_rewards[:, 5].tolist(),
        "frame_max_torque"          : frame_rewards[:, 6].tolist(),
        "frame_GM_velocity"         : frame_rewards[:, 7].tolist(),
        "frame_settling_time_001p"  : frame_rewards[:, 8].tolist(),
        "frame_settling_time_01p"   : frame_rewards[:, 9].tolist(),  
        "frame_settling_time_1p"    : frame_rewards[:, 10].tolist(), 
        "frame_settling_time_3p"    : frame_rewards[:, 11].tolist(),
        "frame_damping_ratio"       : frame_rewards[:, 12].tolist(),
        "frame_Ess_step"            : frame_rewards[:, 13].tolist(),
        "frame_Ess_trap"            : frame_rewards[:, 14].tolist(),
        "frame_overshoot_load"      : frame_rewards[:, 15].tolist(),
        "frame_settling_time_load"  : frame_rewards[:, 16].tolist(), 
        "frame_max_torque_flag"     : frame_rewards[:, 17].tolist(),
        "frame_GM_velocity_cost"    : frame_rewards[:, 18].tolist(),
        "frame_Ess_step"            : frame_rewards[:, 19].tolist(),
        "frame_Ess_trap"            : frame_rewards[:, 20].tolist(),
        "frame_overshoot_load_cost" : frame_rewards[:, 21].tolist(),
        "frame_total_reward"        : frame_rewards[:, 22].tolist()
    }

    network_info = {
        "value_loss": value_losses,
        "policy_loss": policy_losses
    }

    save_result_to_excel(reward_info, results_folder+'reward.xlsx')
    save_result_to_excel(network_info, results_folder+'loss.xlsx')

    for idx, (name, data) in enumerate(reward_info.items()):
        if name == "frame_idx":
            continue
        plt.figure(idx)
        plt.plot(data)
        plt.title(name)
        plt.xlabel("Frame")
        plt.ylabel("Reward")
        # plt.ylim(min(data)-1, max(data)+1)
        plt.axhline(y=0, color='r', linestyle='-')
        plt.autoscale()
        plt.yticks(plt.yticks()[0][::len(plt.yticks()[0])//5])
        plt.savefig(graph_folder + name + '.png')
        plt.savefig(graph_folder + name + '.svg')
        plt.clf()

    plt.figure(idx+1)
    plt.plot(value_losses[1:])
    plt.title("Loss of Value Network")
    plt.xlabel("Frame")
    plt.ylabel("Loss")
    # plt.ylim(min(value_losses[1:])-1, max(value_losses[1:])+1)
    plt.autoscale()
    plt.yticks(plt.yticks()[0][::len(plt.yticks()[0])//5])
    plt.savefig(graph_folder+'value_network_loss.png')
    plt.savefig(graph_folder+'value_network_loss.svg')
    plt.clf()

    plt.figure(idx+2)
    plt.plot(policy_losses[1:])
    plt.title("Loss of Policy Network")
    plt.xlabel("Frame")
    plt.ylabel("Loss")
    plt.ylim(min(policy_losses[1:])-1, max(policy_losses[1:])+1)
    plt.yticks(plt.yticks()[0][::len(plt.yticks()[0])//5])
    plt.savefig(graph_folder+'policy_network_loss.png')
    plt.savefig(graph_folder+'policy_network_loss.svg')
    plt.clf()

def trend_MO_SettlingTime(param, MO_values, ST_valuse):
    # Create the figure and primary y-axis
    fig, ax1 = plt.subplots()

    # Plot the left y-axis data
    ax1.plot(param["values"], MO_values, color='blue', label='Maximum Overshoot')
    ax1.set_xlabel(param["name"])
    ax1.set_ylabel('Maximum Overshoot (%)', color='blue')
    ax1.tick_params(axis='y', labelcolor='blue')

    # Create a secondary y-axis
    ax2 = ax1.twinx()
    ax2.plot(param["values"], ST_valuse, color='red', label='Settling Time')
    ax2.set_ylabel('Settling Time (s)', color='red')
    ax2.tick_params(axis='y', labelcolor='red')

    # Show the plot
    fig.tight_layout()
    plt.show()

def draw_graph_again(file_path):
    reward_info = pd.read_excel(file_path+'reward.xlsx')

    for idx, (name, data) in enumerate(reward_info.items()):
        if name == "frame_idx":
            continue
        plt.figure(idx)
        plt.plot(data)
        plt.title(name)
        plt.xlabel("Frame")
        plt.ylabel("Reward")
        plt.autoscale()
        plt.axhline(y=0, color='r', linestyle='-')
        plt.yticks(plt.yticks()[0][::len(plt.yticks()[0])//5])
        plt.savefig(file_path + 'graph/' + name + '.png')
        plt.savefig(file_path + 'graph/'  + name + '.svg')
        plt.clf()

if __name__ == "__main__":
    # Sample data
    # parameter = {
    #     "name"   : "Kpp",
    #     "values" : [1, 2, 3, 4, 5],
    # }
    # y_left  = [10, 20, 30, 40, 50]  # left y-axis data
    # y_right = [55* 0.001, 45* 0.001, 35* 0.001, 25* 0.001, 15* 0.001] # right y-axis data
    # trend_MO_SettlingTime(parameter, y_left, y_right)


    # data = generate_fake_data()
    # save_result(data, 'training_results.csv')
    # results = load_result('training_results.csv')
    # print(results.__dict__)
    # save_result(results.__dict__, 'training_results_2.csv')

    draw_graph_again('../record/20241115_120325_好的/')
    pass